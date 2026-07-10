import streamlit as st
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from extract_tables import extract_tables
from fill_fisa import fill_fisa

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "_FIsa_Prototip.xlsx")

st.set_page_config(page_title="PDF Table Extractor", page_icon="📄", layout="wide")
st.title("📄 Mengenkalkulation Extractor")

col_stanga, col_dreapta = st.columns(2)

# ── STÂNGA: Date proiect ──────────────────────────────────────
with col_stanga:
    st.subheader("Date proiect")
    client         = st.text_input("Client")
    data           = st.date_input("Data")
    tip_proiect    = st.selectbox("Tip proiect", ["", "Mobilier în Roomdesigner", "Debitare căntuire"])
    tip_solicitare = st.selectbox("Tip solicitare", ["", "Ofertă", "Comandă (cu work preparation)"])
    preluat_de     = st.selectbox("Preluat de", ["", "Dan Slotea", "David Constantin", "Iulian Necula",
                                                  "Liliana Chiriță", "Adrian Mărgărit", "Denisa Manea"])
    proiectat_de   = st.selectbox("Proiectat de", ["", "Client", "Plan M"])

    st.divider()
    uploaded_file = st.file_uploader("Încarcă PDF-ul", type="pdf")

# ── DREAPTA: Accesorii ────────────────────────────────────────
with col_dreapta:
    accesorii = []

    # ── Expander 1: Accesorii ─────────────────────────────────
    with st.expander("Accesorii", expanded=False):
        # VB35/Cabineo Culori
        c1, c2 = st.columns([1, 2])
        with c1:
            st.markdown("VB35/Cabineo Culori")
        with c2:
            val_vb = st.text_input("", key="val_vb", label_visibility="collapsed", placeholder="Detalii")
            if val_vb:
                accesorii.append(("VB35/Cabineo Culori", val_vb))

        # Alte Accesorii — câmpuri dinamice cu +
        st.markdown("**Alte Accesorii**")
        if "alte_acc_count" not in st.session_state:
            st.session_state.alte_acc_count = 0
        for i in range(st.session_state.alte_acc_count):
            val = st.text_input("", key=f"alte_acc_{i}", label_visibility="collapsed",
                                placeholder=f"Accesoriu #{i+1}")
            accesorii.append((f"Alte Accesorii #{i+1}" if st.session_state.alte_acc_count > 1 else "Alte Accesorii", val))
        if st.button("+ Adaugă accesoriu"):
            st.session_state.alte_acc_count += 1
            st.rerun()

    # ── Expander 2: GOLA ──────────────────────────────────────
    with st.expander("GOLA", expanded=False):
        GOLA_OPTIUNI = [
            "Culoare",
            "L orizontal",
            "C orizontal",
            "L vertical",
            "C vertical",
            "Prinderi L",
            "Prinderi C",
            "Capace L, perechi",
            "Capace C, perechi",
            "Prinderi colț interior",
            "Prinderi colț exterior",
        ]
        for optiune in GOLA_OPTIUNI:
            c1, c2 = st.columns([1, 2])
            with c1:
                st.markdown(optiune)
            with c2:
                val = st.text_input("", key=f"val_{optiune}",
                                    label_visibility="collapsed",
                                    placeholder="Detalii")
                if val:
                    accesorii.append((optiune, val))

    # ── Expander 3: Rame Aluminiu ─────────────────────────────
    with st.expander("Rame Aluminiu", expanded=False):
        if "rama_count" not in st.session_state:
            st.session_state.rama_count = 0
        for i in range(st.session_state.rama_count):
            val = st.text_input("", key=f"rama_{i}", label_visibility="collapsed",
                                placeholder=f"Ramă aluminiu #{i+1}")
            label = f"Ramă aluminiu #{i+1}" if st.session_state.rama_count > 1 else "Ramă aluminiu"
            accesorii.append((label, val))
        if st.button("+ Adaugă ramă"):
            st.session_state.rama_count += 1
            st.rerun()

# ── Extragere ─────────────────────────────────────────────────
antet = {
    "B3": client,
    "B6": str(data),
    "I3": tip_proiect,
    "I4": tip_solicitare,
    "I5": preluat_de,
    "I6": proiectat_de,
}

if uploaded_file:
    st.success(f"Fișier încărcat: {uploaded_file.name}")
    filename_stem = os.path.splitext(uploaded_file.name)[0]

    if st.button("Extrage"):
        with st.spinner("Se extrag datele..."):
            tmp_pdf = "/tmp/input.pdf"
            with open(tmp_pdf, "wb") as f:
                f.write(uploaded_file.read())
            tables = extract_tables(tmp_pdf, "/tmp")
        meta = tables.get("_metadata", {})

        st.success("Extragere finalizată!")
        if meta.get("nr_proiect"):
            st.info(f"Detectat automat: Nr. proiect **{meta['nr_proiect']}** — {meta['nume_proiect']}")

        # ── Avertizari ────────────────────────────────────────
        campuri_lipsa = []
        if not client: campuri_lipsa.append("Client")
        if not tip_proiect: campuri_lipsa.append("Tip proiect")
        if not tip_solicitare: campuri_lipsa.append("Tip solicitare")
        if not preluat_de: campuri_lipsa.append("Preluat de")
        if not proiectat_de: campuri_lipsa.append("Proiectat de")
        if campuri_lipsa:
            st.warning(f"⚠️ Date proiect incomplete: {', '.join(campuri_lipsa)}")

        if not accesorii:
            st.warning("⚠️ Dl. Dan te rog sa adaugi accesorii la acest proiect.")

        if os.path.exists(TEMPLATE_PATH):
            with st.spinner("Se completează fișa..."):
                out_fisa = f"/tmp/{filename_stem}_fisa.xlsx"
                fill_fisa(tmp_pdf, out_fisa, TEMPLATE_PATH, antet=antet, meta=meta, accesorii=accesorii)
            with open(out_fisa, "rb") as f:
                st.download_button(
                    label="⬇️ Descarcă Fișa Completată",
                    data=f.read(),
                    file_name=f"{meta.get('nr_proiect','')}_{meta.get('nume_proiect','')}_{client}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("Template _FIsa_Prototip.xlsx nu a fost găsit în repository.")

        with st.spinner("Se generează Excel..."):
            wb = Workbook()
            wb.remove(wb.active)
            header_font = Font(bold=True, name="Arial")
            header_fill = PatternFill("solid", start_color="D9E1F2")

            for sheet_name, rows in tables.items():
                if sheet_name == "_metadata":
                    continue
                ws = wb.create_sheet(title=sheet_name[:31])
                for r_idx, row in enumerate(rows, start=1):
                    full_row = ["Source File"] + row if r_idx == 1 else [filename_stem] + row
                    for c_idx, value in enumerate(full_row, start=1):
                        if r_idx > 1:
                            try:
                                num = float(value)
                                value = int(num) if num == int(num) else num
                            except (ValueError, TypeError):
                                pass
                        cell = ws.cell(row=r_idx, column=c_idx, value=value)
                        if r_idx == 1:
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = Alignment(horizontal="center")
                for col in ws.columns:
                    max_len = max((len(str(c.value)) for c in col if c.value), default=10)
                    ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            xlsx_buffer = io.BytesIO()
            wb.save(xlsx_buffer)
            xlsx_buffer.seek(0)

        st.download_button(
            label="⬇️ Descarcă Tabele Excel (5 sheet-uri)",
            data=xlsx_buffer,
            file_name=f"{filename_stem}_tabele.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
