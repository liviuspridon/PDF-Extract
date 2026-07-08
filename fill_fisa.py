#!/usr/bin/env python3
"""
Ia un PDF Mengenkalkulation si completeaza _FIsa_Prototip.xlsx
cu datele extrase, pastrand toate formulele intacte.

Scriere:
  Deco  : A=Material, B=Grosime, C=Cant, D=Lung, E=Lat
  Cant  : A=Denumire, B=Inaltime, C=Bucata, D=Lungime, E=Net_m
  Taiere: A=Material, B=Grosime, C=metri, D=metri_per_qm

Usage: python3 fill_fisa.py input.pdf [output.xlsx]
"""

import sys
import os
import shutil
from openpyxl import load_workbook
from extract_tables import extract_tables

TEMPLATE = os.path.join(os.path.dirname(__file__), "_FIsa_Prototip.xlsx")


def fill_fisa(pdf_path, out_path=None, template_path=None, antet=None, meta=None):
    if template_path is None:
        template_path = TEMPLATE

    if out_path is None:
        base = os.path.splitext(os.path.basename(pdf_path))[0]
        out_path = f"{base}_fisa.xlsx"

    # extract tables from PDF
    tables = extract_tables(pdf_path, "/tmp")

    # copy template — never modify the original
    shutil.copy2(template_path, out_path)
    wb = load_workbook(out_path)

    # ── ANTET (date proiect) ─────────────────────────────────────
    ws_fisa = wb.worksheets[0]  # primul sheet = Fisa

    # ── Metadata din PDF (nr proiect, titlu) ────────────────────
    # B4:D4 si B5:D5 sunt merged cells — scriem in celula de start
    if meta:
        if meta.get("nr_proiect"):
            ws_fisa.cell(row=5, column=2).value = str(meta["nr_proiect"])
            ws_fisa.cell(row=5, column=2).number_format = "@"
        if meta.get("nume_proiect"):
            ws_fisa.cell(row=4, column=2).value = meta["nume_proiect"]

    # ── Antet (campuri din formular) ─────────────────────────────
    if antet:
        for cell_ref, value in antet.items():
            if value:  # scrie doar daca nu e gol
                if cell_ref == "B5":
                    cell = ws_fisa[cell_ref]
                    cell.value = str(value)
                    cell.number_format = "@"
                else:
                    ws_fisa[cell_ref] = value

    # ── DECO (from Plattenmaterial) ──────────────────────────────
    # PDF headers: Pos, Material, Starke, Stuck, Lange, Breite, ...
    plat = tables.get("Plattenmaterial", [])
    if len(plat) > 1:
        ws = wb["Deco"]
        for i, row in enumerate(plat[1:], start=2):  # skip header
            # row = [Pos, Material, Starke, Stuck, Lange, Breite, ...]
            _, material, starke, stuck, lange, breite = row[0], row[1], row[2], row[3], row[4], row[5]
            vk_qm = row[16] if len(row) > 16 else None
            summe = row[17] if len(row) > 17 else None
            ws.cell(row=i, column=1).value = material       # A
            ws.cell(row=i, column=2).value = _num(starke)   # B Grosime
            ws.cell(row=i, column=3).value = _num(stuck)    # C Cant
            ws.cell(row=i, column=4).value = _num(lange)    # D Lung
            ws.cell(row=i, column=5).value = _num(breite)   # E Lat
            ws.cell(row=i, column=16).value = _num(vk_qm)  # P VK qm
            ws.cell(row=i, column=17).value = _num(summe)   # Q Summe
        print(f"Deco: wrote {len(plat)-1} rows")

    # ── CANT (from Material pentru margini) ──────────────────────
    # PDF headers: Pos, Material, Inaltime, Bucata, Lungime, Net_m, ...
    marg = tables.get("Material_pentru_margini", [])
    if len(marg) > 1:
        ws = wb["Cant"]
        for i, row in enumerate(marg[1:], start=2):
            _, material, inaltime, bucata, lungime, net_m = row[0], row[1], row[2], row[3], row[4], row[5]
            ws.cell(row=i, column=1).value = material        # A Denumire
            ws.cell(row=i, column=2).value = _num(inaltime)  # B Inaltime
            ws.cell(row=i, column=3).value = _num(bucata)    # C Cant
            ws.cell(row=i, column=4).value = _num(lungime)   # D Lung
            ws.cell(row=i, column=5).value = _num(net_m)     # E Net_m
            # H = RIGHT(A,3) = ultimele 3 caractere din denumire (ex: "2.0", "0.8")
            # scris explicit ca sa nu depinda de ordinea recalcularii formulelor
            ws.cell(row=i, column=8).value = material[-3:] if len(material) >= 3 else material
        print(f"Cant: wrote {len(marg)-1} rows")

    # ── TAIERE (from Taiere) ─────────────────────────────────────
    # PDF headers: Pos, Material, Grosime, metri, metri_per_qm, ...
    taiere = tables.get("Taiere", [])
    if len(taiere) > 1:
        ws = wb["Taiere"]
        for i, row in enumerate(taiere[1:], start=2):
            _, material, grosime, metri, metri_per_qm = row[0], row[1], row[2], row[3], row[4]
            ws.cell(row=i, column=1).value = material           # A
            ws.cell(row=i, column=2).value = _num(grosime)      # B Grosime
            ws.cell(row=i, column=3).value = _num(metri)        # C metri
            ws.cell(row=i, column=4).value = _num(metri_per_qm) # D metri/qm
        print(f"Taiere: wrote {len(taiere)-1} rows")

    # tell Excel to recalculate everything on open
    wb.calculation.calcMode = "auto"

    wb.save(out_path)
    print(f"Saved: {out_path}")
    return out_path


def _num(val):
    """Convert string to int or float, leave as-is if not numeric."""
    try:
        f = float(str(val).replace(",", "."))
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return val


if __name__ == "__main__":
    pdf_path = sys.argv[1] if len(sys.argv) > 1 else "input.pdf"
    out_path = sys.argv[2] if len(sys.argv) > 2 else None
    fill_fisa(pdf_path, out_path)
